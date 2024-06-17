import pandas as pd
from openpyxl import load_workbook
import numpy as np
import tkinter as tk
from tkinter import filedialog, ttk

path=""

def Fehlermeldung():
    Fehler = tk.Tk()
    Fehler.title("Fehler beim Öffnen der Datei")
    Fehler.geometry("500x100")
    Fehlertxt = tk.Label(Fehler, text="Bitte geben Sie den Dateityp .xlsx ein.")
    okbutt= tk.Button(Fehler, text ="OK", command=Fehler.destroy)
    Fehlertxt.pack()
    okbutt.pack()
    Fehler.mainloop()

def Hilfetext():
    # Hauptfenster erstellen
    Hilfetxt = tk.Tk()
    Hilfetxt.title("Scrollbar Beispiel")
    Hilfetxt.geometry("900x900")

    # Erstelle eine vertikale Scrollbar
    scrollverti = ttk.Scrollbar(Hilfetxt, orient="vertical")
    scrollverti.pack(side="right", fill="y")

    # Erstelle eine horizontale Scrollbar
    scrollbar_hor = ttk.Scrollbar(Hilfetxt, orient="horizontal")
    scrollbar_hor.pack(side="bottom", fill="x")

    # Erstelle einen Textbereich, dem die Scrollbars hinzugefügt werden
    text_area = tk.Text(Hilfetxt, wrap="none", yscrollcommand=scrollverti.set, xscrollcommand=scrollbar_hor.set)
    text_area.pack(side="left", fill="both", expand=True)

    # Verbinde die Scrollbars mit dem Textbereich
    scrollverti.config(command=text_area.yview)
    scrollbar_hor.config(command=text_area.xview)

    # Öffne die Datei und lese den Inhalt im UTF-8-Format
    with open('data/hilftxt', 'r', encoding='utf-8') as file:
        data = file.read()

    # Füge den Inhalt der Datei in den Textbereich ein
    text_area.insert("end", data)

    # Setze den Textbereich in den nur-Lese-Modus
    text_area.config(state="disabled")
    Hilfetxt.mainloop()

#from openpyxl.utils import get_column_letter

def get_file_path():
    global path
    # Dateidialog öffnen, um einen Dateipfad auszuwählen
    path = filedialog.askopenfilename()

    # Den ausgewählten  Dateipfad anzeigen
    if path:
        path_label.config(text="Dateipfad: " + path)
    else:
        path_label.config(text="Keine Datei ausgewählt")


def Hauptfenster():
    root = tk.Tk()
    root.title("Dateipfad auswählen")
    root.state('zoomed')
    def ren ():
        if check_path() == True:
            root.destroy()
    # Button zum Öffnen des Dateidialogs
    browse_button = tk.Button(root, text="Datei auswählen", command=get_file_path)
    browse_button2 = tk.Button(root, text="Bestätigen", command= ren)
    browse_hilfe = tk.Button(root, text="Hilfe", command=Hilfetext, bg="#FF4500")
    browse_hilfe.place(relx=1.0, rely=0.0, anchor='ne', x=-10, y=10)
    browse_button.pack(anchor='n', pady=10)
    browse_button2.pack(pady=10)



    # Label zur Anzeige des ausgewählten Dateipfads
    global path_label
    path_label = tk.Label(root, text="Dateipfad: ")
    path_label.pack()


    root.mainloop()

def read_clean_val(messungen, start_list, end_list, Data):

   i = 0
   data = []
   while i < messungen:
      start = start_list[i]
      end = end_list[i]
      #print(Data_x[start:end])
      data.append(Data[start:end])
      i += 1
   #print(Data_x)
   return data


def get_range():
   start = []
   end = []
   active = False
   for cell in ws['A']:
      if cell.value == 'Frame':
         start.append(cell.row)
         active = True
      #print(cell.value)
      if active and cell.value is None:
         end.append(cell.row-1)
         active = False
   return (start, end)


def get_data():
   Data_x = []
   Data_y = []
   for cell in ws['C']:
      Data_x.append(cell.value)
   for cell in ws['D']:
      Data_y.append(cell.value)
   return Data_x, Data_y

def get_df(data):
    h= 0
    minLen=len(data[0])
    for h in range(len(data)):
        if len(data[h])<=minLen:
            minLen=len(data[h])
        h+=1

    for k in range(len(data)):
        data[k]=np.array(data[k][:minLen])
    #print(frame)
    #print(minLen)

    df=pd.DataFrame(
        {
            'LinFus1': data[0],
            'LinFus2': data[1],
            'LinFus3': data[2],
            'RecFus1': data[3],
            'RecFus2': data[4],
            'RecFus3': data[5],
        }
    )
    return df

def Fus_mit(df_in):
    begin = round(len(df_in) * (0.05)/2)
    back = round(len(df_in)-begin)

    #df_sort = df_in.apply(lambda x: x.sort_values().values)
    df_sort = pd.DataFrame(np.sort(df_in.values, axis=0), index=df_in.index, columns=df_in.columns)
    df_out = df_sort.take([begin, back])
    #min_max = df_out['LinFus1'].tolist()
    min_max = []
    for column in df_out.columns:
        li = df_out[column].tolist()
        min_max.append(li)
    dist = []
    for i in range(len(min_max)):
        if min_max[i][0] >= min_max[i][1] :
            dist.append(round(min_max[i][0] - min_max[i][1], 2))
        else:
            dist.append(round(min_max[i][1] - min_max[i][0], 2))

    mid_val =[round((dist[0] + dist[1] + dist[2])/3, 2), round((dist[3] + dist[4] + dist[5])/3, 2)]
    print(dist)

    return df_out, mid_val

def check_path():
    global path
    if ".xlsx" not in path:
        Fehlermeldung()
        return "NotX"
    else :
        return True

def auswerutng() :
    #if check_path() != "NotX":
      #  wb = load_workbook(path)
      #  ws = wb.active
        #pop Up fenster
        datenausertung = tk.Tk()
        datenausertung.title("Auswertung")
        datenausertung.state('zoomed')

        Data_x, Data_y = get_data()
        start_list, end_list = get_range()
        end_list.append(len(ws['c']))
        messungen = (len(start_list))
        print("Anzahl der Messungen: ", messungen)

        df_x = get_df(read_clean_val(messungen, start_list, end_list, Data_x))
        df_y = get_df(read_clean_val(messungen, start_list, end_list, Data_y))
        #print(df_x)
        #print(df_y)
        df_x_sort, mid_val_x = Fus_mit(df_x)
        df_y_sort, mid_val_y = Fus_mit(df_y)
        print(df_x_sort)
        print(df_y_sort)

        print(mid_val_x)
        print(mid_val_y)

        #ausgabe des data Frames
        daten_label = tk.Label(datenausertung, text="" + df_x_sort.to_string() + "\n\n Ovalerdurchmesser Links (X in mm):" + str(mid_val_x[0]) + "\t\tOvalerdurchmesser Links (Y in mm):" + str(mid_val_y[0]) + "\n\nOvalerdurchmesser Rechts (X in mm):" + str(mid_val_x[1]) + "\t\tOvalerdurchmesser Rechts (Y in mm):" + str(mid_val_y[1]))
        daten_label.pack()
        datenausertung.mainloop()





Hauptfenster()
wb = load_workbook(path)
ws = wb.active
auswerutng()





