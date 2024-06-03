import sys

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import numpy as np
import tkinter as tk
from tkinter import filedialog
import os
import asposecells


#from openpyxl.utils import get_column_letter
path = ""
def get_file_path():
    global path
    # Dateidialog öffnen, um einen Dateipfad auszuwählen
    path = filedialog.askopenfilename()

    # Den ausgewählten Dateipfad anzeigen
    if path:
        path_label.config(text="Dateipfad: " + path)
    else:
        path_label.config(text="Keine Datei ausgewählt")


def Hauptfenster():
    root = tk.Tk()
    root.title("Dateipfad auswählen")
    root.geometry("900x400")
    # Button zum Öffnen des Dateidialogs
    browse_button = tk.Button(root, text="Datei auswählen", command=get_file_path)
    browse_button2 = tk.Button(root, text="Bestätigen", command=root.destroy)
    browse_button.pack(pady=10)
    browse_button2.pack(pady=10)

    # Label zur Anzeige des ausgewählten Dateipfads
    global path_label
    path_label = tk.Label(root, text="Dateipfad: ")
    path_label.pack()

    root.mainloop()

def read_clean_val(messungen, start_list, end_list, Data):

   i = 0
   data = []
   while i < messungen:
      start = start_list[i]
      end = end_list[i]
      #print(Data_x[start:end])
      data.append(Data[start:end])
      i += 1
   #print(Data_x)
   return data


def get_range():
   start = []
   end = []
   active = False
   for cell in ws['A']:
      if cell.value == 'Frame':
         start.append(cell.row)
         active = True
      #print(cell.value)
      if active and cell.value is None:
         end.append(cell.row-1)
         active = False
   return (start, end)


def get_data():
   Data_x = []
   Data_y = []
   for cell in ws['C']:
      Data_x.append(cell.value)
   for cell in ws['D']:
      Data_y.append(cell.value)
   return Data_x, Data_y

def get_df(data):
    h= 0
    minLen=len(data[0])
    for h in range(len(data)):
        if len(data[h])<=minLen:
            minLen=len(data[h])
        h+=1

    for k in range(len(data)):
        data[k] = np.array(data[k][:minLen])

    df = pd.DataFrame(
        {
            'LinFus1': data[0],
            'LinFus2': data[1],
            'LinFus3': data[2],
            'RecFus1': data[3],
            'RecFus2': data[4],
            'RecFus3': data[5],
        }
    )
    return df

def check_path():
    global path
    if ".xls"  in path and ".xlsx" not in path:
        print("Keine Valide Datei")
        return "NotX"

def auswerutng() :
    #if check_path() != "NotX":
      #  wb = load_workbook(path)
      #  ws = wb.active
        Data_x, Data_y = get_data()
        start_list, end_list = get_range()
        end_list.append(len(ws['c']))
        messungen = (len(start_list))

        print("Anzahl der Messungen: ", messungen)
        print(start_list)
        print(end_list)

        data_x = read_clean_val(messungen, start_list, end_list, Data_x)
        data_y = read_clean_val(messungen, start_list, end_list, Data_y)
        #print(data_x)
        #print(data_x[2])
        #print(Data_y)
        print(data_y[2])
        print(data_x[2])

        df_x = get_df(data_x)
        df_y = get_df(data_y)

        print(df_x)
        print(df_y)




Hauptfenster()
if check_path() != "NotX":
    wb = load_workbook(path)
    ws = wb.active
    auswerutng()
