import pathlib
from typing import List, Tuple, Set

import matplotlib.axes
import pandas as pd
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from matplotlib.patches import Ellipse
import matplotlib.transforms as transforms

path = pathlib.Path('./data')


# Some Simple wrapper classes to verbalize structure

class MeasurementRange:
    def __init__(self, start: int, end: int):
        self.start = start
        self.end = end
        self.df: pd.DataFrame = None

    def __repr__(self):
        return f'({self.start} ; {self.end})'


class Dataset:
    def __init__(self, filename: pathlib.Path, sheet: str):
        self.filename: pathlib.Path = filename
        self.sheet_name: str = sheet
        self.measurement_ranges: List[MeasurementRange] = []
        self.data_frame: pd.DataFrame = pd.DataFrame()
        self.max_rows: int = 0
        self.plot: matplotlib.axes.Axes = None

    def __repr__(self):
        return (str(self.filename)
                + ':'
                + self.sheet_name
                + ', '.join([str(f) for f in self.measurement_ranges]))


def calc_avg(df: pd.DataFrame, destination: str, field: str, count: int):
    df[destination] = df.apply(
        lambda row: sum([row[f'{field}{i}'] for i in range(count)]) / count,
        axis=1)


def calc_axes(df: pd.DataFrame, column: str, padding: int = 10) -> tuple[int, int]:
    v_max: int = df[column].max()
    v_min: int = df[column].min()
    if pd.isna(v_max):
        v_max = 10
    if pd.isna(v_min):
        v_min = 10

    value: int = max(abs(v_min), abs(v_max))
    padded: int = value + padding
    print(column, v_min, v_max, 'Absolut:', value, 'Padded:', padded)

    return padded * -1, padded


def scan_file(file: pathlib.Path) -> Dataset:
    wb = load_workbook(file, False, False, False, False, False)
    dataset = Dataset(file, wb.sheetnames[0])
    ws = wb[dataset.sheet_name]
    r = MeasurementRange(0, 0)
    active = False
    for cell in ws['A']:
        if cell.value == 'Frame':
            r.start = cell.row
            active = True
        if active and cell.value is None:
            r.end = cell.row
            dataset.measurement_ranges.append(r)
            r = MeasurementRange(0, 0)
            active = False

    dataset.max_rows = max([r.end - r.start for r in dataset.measurement_ranges])
    return dataset


def load_ranges(dataset: Dataset):
    for r in dataset.measurement_ranges:
        df = pd.read_excel(str(dataset.filename),
                           sheet_name=dataset.sheet_name,
                           header=r.start - 1,
                           index_col=0,
                           usecols='A:E',
                           nrows=r.end - r.start)
        r.df = df


def populate_common_df(dataset: Dataset):
    # Merge dataframes
    dataset.data_frame.index.name = 'Frame'  # Set name of index column
    dataset.data_frame['t'] = sorted([f.df for f in dataset.measurement_ranges], key=len)[-1]['ms']
    for index, r in enumerate(dataset.measurement_ranges):
        dataset.data_frame[f'x{index}'] = r.df['X (mm)']
        dataset.data_frame[f'y{index}'] = r.df['Y (mm)']
        dataset.data_frame[f'f{index}'] = r.df[f'Force (N)']

    # Calculate averages for x, y and force
    calc_avg(dataset.data_frame, 'avg_x', 'x', len(dataset.measurement_ranges))
    calc_avg(dataset.data_frame, 'avg_y', 'y', len(dataset.measurement_ranges))
    calc_avg(dataset.data_frame, 'avg_f', 'f', len(dataset.measurement_ranges))


def set_axes(dataset: Dataset):
    # Get range of x
    x_min, x_max = calc_axes(dataset.data_frame, 'avg_x')

    # Get range of y
    y_min, y_max = calc_axes(dataset.data_frame, 'avg_y')

    plt.xlim(x_min, x_max)
    plt.ylim(y_min, y_max)


def main():
    datasets: List[Dataset] = []  # This contains all files

    # Iterate over all files and scan datasets.
    for file in [f for f in path.iterdir() if f.is_file() and f.name.endswith('.xlsx')]:
        datasets.append(scan_file(file))

    # Load .gitkeep-frames for every range
    for dataset in datasets:
        load_ranges(dataset)

        populate_common_df(dataset)

        # Kill all rows that contain empty values.
        dataset.data_frame.dropna(axis=0, inplace=True)

        print(dataset)
        print(dataset.data_frame)
        dataset.plot = dataset.data_frame.plot.scatter(x='avg_x', y='avg_y', c='avg_f')

        # Add Axis lines
        dataset.plot.axhline(c='grey', lw=1)
        dataset.plot.axvline(c='grey', lw=1)

        set_axes(dataset)

        # Draw Ellipse over plot
        ellipse = Ellipse((0, 0), 5, 8, facecolor='none', edgecolor='red')
        dataset.plot.add_patch(ellipse)

        # Show plot
        plt.show()


if __name__ == '__main__':
    main()
